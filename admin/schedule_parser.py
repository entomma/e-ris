#!/usr/bin/env python3
"""
schedule_parser.py
Parses the official Pampanga State University class schedule XLSX and outputs JSON.
"""

import sys
import json
import re
import openpyxl

DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']

def get_val(cell):
    return str(cell.value).strip() if cell and cell.value is not None else ''

def parse_time(raw):
    """Convert '7:00', '7:30', '1:00' → 'HH:MM:SS' with smart PM detection."""
    if not raw:
        return None
    raw = str(raw).strip()
    
    # Clean up whitespace and make lowercase
    raw_lower = re.sub(r'\s+', '', raw).lower()
    try:
        is_pm = 'pm' in raw_lower
        if ':' in raw_lower:
            parts = raw_lower.replace('am', '').replace('pm', '').split(':')
            h, m = int(parts[0]), int(parts[1])
        else:
            h = int(raw_lower.replace('am', '').replace('pm', ''))
            m = 0
        
        # SMART PM INFERENCE:
        # 1. Explicit PM
        if is_pm and h < 12:
            h += 12
        # 2. No marker, but hour is 1-6 (afternoon school hours)
        elif not is_pm and not 'am' in raw_lower:
            if 1 <= h <= 6:
                h += 12
            
        return f"{h:02d}:{m:02d}:00"
    except:
        return None

def parse_time_range(cell_text):
    if not cell_text: return None, None
    s = str(cell_text).strip()
    s = re.sub(r'\s*[-–]\s*', '-', s)
    parts = s.split('-', 1)
    if len(parts) != 2: return None, None
    return parse_time(parts[0].strip()), parse_time(parts[1].strip())

def normalise_course_code(code):
    """Formats everything to 'CC 123(C)' style to ensure database matches perfectly."""
    if not code: return ""
    s = str(code).upper().strip()
    
    # 1. Remove Lec/Lab/Lecture/Laboratory suffixes
    s = re.sub(r'\s*(LEC|LAB|LECTURE|LABORATORY)\s*$', '', s)
    
    # 2. Ensure exactly one space between letters and numbers (e.g. CC123 -> CC 123)
    s = re.sub(r'([A-Z])\s*(\d)', r'\1 \2', s)
    
    # 3. Remove space before parenthesis (e.g. CC 123 (C) -> CC 123(C))
    s = re.sub(r'\s+\(', '(', s)
    
    # 4. Collapse any other double spaces
    s = re.sub(r'\s+', ' ', s)
    
    return s.strip()

def infer_year_level(section_name):
    m = re.search(r'(\d)', str(section_name))
    return int(m.group(1)) if m else 1

def infer_major(section_name):
    m = re.match(r'([A-Z]+)', str(section_name).strip())
    return m.group(1) if m else 'BSIT'

def is_course_code(val):
    if not val: return False
    v = str(val).strip().lower()
    
    if v == 'ojt': return True
    
    if not re.search(r'\d', v): return False
    if not re.match(r'^[a-z]', v): return False
    
    if re.match(r'^(rm|cl)\s*\d+', v): return False
        
    skip = ['republic', 'commission', 'pampanga', 'campus', 'second', 'first',
            'prepared', 'noted', 'approved', 'recommending', 'course ',
            'lec/', 'lab', 'room', 'tba']
    if any(v.startswith(s) for s in skip):
        return False
    return True

def parse_sheet(ws):
    rows = list(ws.iter_rows())

    section_name = ws.title
    for row in rows[:15]:
        v0 = get_val(row[0]).lower()
        if v0.startswith('section'):
            v1 = get_val(row[1])
            if v1: section_name = v1
            break

    year_level = infer_year_level(section_name)
    major      = infer_major(section_name)

    day_row_idx  = None
    day_col_map  = {}

    for i, row in enumerate(rows):
        row_vals = [get_val(c) for c in row]
        if 'Monday' in row_vals:
            day_row_idx = i
            for j, v in enumerate(row_vals):
                if v in DAYS:
                    day_col_map[v] = j
            break

    if day_row_idx is None:
        return None

    time_rows = []
    for i in range(day_row_idx + 1, len(rows)):
        cell_a = get_val(rows[i][0])
        start, end = parse_time_range(cell_a)
        if start:
            time_rows.append((i, start, end))

    if not time_rows:
        return None

    merged_map = {}
    for m_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = m_range.bounds
        for r in range(min_row - 1, max_row):
            for c in range(min_col - 1, max_col):
                merged_map[(r, c)] = (min_row - 1, max_row - 1)

    entries = []

    for day, dcol in day_col_map.items():
        active_block = None

        for row_iter_idx, (tr_idx, start, end) in enumerate(time_rows):
            cell = rows[tr_idx][dcol]
            val = get_val(cell)

            if active_block is None:
                if is_course_code(val):
                    active_block = {
                        'start_time': start,
                        'end_time': end,
                        'texts': [val]
                    }
            else:
                if val:
                    active_block['texts'].append(val)
                active_block['end_time'] = end

            if active_block is not None:
                block_ends = False
                
                # 1. Reliable Check: Excel Merged Cells
                if (tr_idx, dcol) in merged_map:
                    _, max_row = merged_map[(tr_idx, dcol)]
                    if tr_idx == max_row:
                        block_ends = True
                else:
                    # 2. Smart Check: Unmerged Consecutive Blocks
                    if row_iter_idx == len(time_rows) - 1:
                        block_ends = True
                    else:
                        next_r = time_rows[row_iter_idx + 1][0]
                        next_cell = rows[next_r][dcol]
                        next_val = get_val(next_cell)
                        
                        if next_val:
                            # End block immediately if the text is a brand NEW course
                            if is_course_code(next_val):
                                block_ends = True
                            else:
                                block_ends = False
                        else:
                            # Next cell is blank. Do we continue extending?
                            curr_color = cell.fill.start_color.index if cell.fill and cell.fill.start_color else None
                            next_color = next_cell.fill.start_color.index if next_cell.fill and next_cell.fill.start_color else None
                            
                            has_fill = curr_color and curr_color not in ('00000000', 'FFFFFFFF', 0, '0')
                            
                            if has_fill and curr_color == next_color:
                                block_ends = False
                            else:
                                block_ends = True

                if block_ends:
                    texts = active_block['texts']
                    course_code = texts[0]
                    room = texts[-1] if len(texts) > 1 else 'TBA'
                    instructor = " / ".join(texts[1:-1]) if len(texts) > 2 else (texts[1] if len(texts) == 2 else '')
                    
                    entries.append({
                        'day'        : day,
                        'start_time' : active_block['start_time'],
                        'end_time'   : active_block['end_time'],
                        'course_code': normalise_course_code(course_code),
                        'instructor' : instructor,
                        'room'       : room,
                    })
                    active_block = None

    # Redundancy Check: Seamlessly merge accidental splits of the exact same subject
    merged = {}
    for e in entries:
        key = (e['day'], e['course_code'])
        if key not in merged:
            merged[key] = dict(e)
        else:
            prev_end = merged[key]['end_time']
            cur_start = e['start_time']
            if cur_start == prev_end:
                merged[key]['end_time'] = e['end_time']

    schedule_entries = list(merged.values())

    # Bottom Sheet Course List Scanner
    course_list = []
    for i in range(day_row_idx + 1, len(rows)):
        row = rows[i]
        code = get_val(row[3]) if len(row) > 3 else ''
        title = get_val(row[4]) if len(row) > 4 else ''
        units_raw = row[6].value if len(row) > 6 else None

        if not code or not title: continue
        if code.lower() in ('course code', 'code'): continue
        if not re.search(r'\d', code): continue
        if not re.match(r'^[A-Za-z]', code): continue

        lec_units = 3
        if units_raw is not None:
            u = str(units_raw).strip()
            m = re.match(r'(\d+)', u)
            if m: lec_units = float(m.group(1))

        course_list.append({
            'code' : normalise_course_code(code),
            'title': title,
            'units': lec_units,
        })

    return {
        'section'         : section_name,
        'major'           : major,
        'year_level'      : year_level,
        'schedule_entries': schedule_entries,
        'course_list'     : course_list,
    }


def main():
    if len(sys.argv) < 2:
        print(json.dumps({'error': 'No file path provided'}))
        sys.exit(1)

    path = sys.argv[1]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as ex:
        print(json.dumps({'error': str(ex)}))
        sys.exit(1)

    results = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parsed = parse_sheet(ws)
        if parsed and parsed['schedule_entries']:
            results.append(parsed)

    print(json.dumps(results, ensure_ascii=False, indent=2))

if __name__ == '__main__':
    main()